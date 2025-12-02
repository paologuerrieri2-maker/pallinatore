#!/usr/bin/env python3
"""
Pallinatore Quote v6
- Pallini trascinabili con mouse
- Eliminazione con click destro
- Rinumerazione automatica
- Barra progressiva durante OCR
"""

import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from PIL import Image, ImageTk
import numpy as np

# PaddleOCR - inizializzazione lazy
_ocr_engine = None

def get_ocr_engine():
    global _ocr_engine
    if _ocr_engine is None:
        from paddleocr import PaddleOCR
        _ocr_engine = PaddleOCR(lang="en", show_log=False)
    return _ocr_engine


def run_ocr(image_path, progress_callback=None):
    """Esegue OCR su un'immagine."""
    ocr = get_ocr_engine()
    
    with Image.open(image_path) as img:
        orig_w, orig_h = img.size
    
    if progress_callback:
        progress_callback(10, "Analisi immagine...")
    
    raw = ocr.ocr(image_path)
    
    if progress_callback:
        progress_callback(70, "Elaborazione risultati...")
    
    if raw is None or len(raw) == 0:
        return []
    
    results = []
    res_obj = raw[0]
    
    try:
        polys = res_obj.get('rec_polys') or res_obj.get('dt_polys')
        texts = res_obj.get('rec_texts', [])
        scores = res_obj.get('rec_scores', [])
        
        if polys is None or len(polys) == 0:
            return []
        
        total = len(polys)
        for i, poly in enumerate(polys):
            box = np.array(poly, dtype=float)
            text = texts[i] if i < len(texts) else ""
            conf = float(scores[i]) if i < len(scores) and scores[i] is not None else 0.0
            results.append({"box": box, "text": str(text), "conf": conf})
            
            if progress_callback and i % 10 == 0:
                pct = 70 + int(25 * i / total)
                progress_callback(pct, f"Elaborazione {i+1}/{total}...")
    
    except Exception as e:
        print(f"[OCR Error] {e}")
    
    if progress_callback:
        progress_callback(100, "Completato!")
    
    return results


def pdf_to_image(pdf_path, dpi=150):
    """Converte PDF in immagine."""
    try:
        import fitz
        doc = fitz.open(pdf_path)
        page = doc[0]
        mat = fitz.Matrix(dpi/72, dpi/72)
        pix = page.get_pixmap(matrix=mat)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        doc.close()
        return img
    except ImportError:
        from pdf2image import convert_from_path
        images = convert_from_path(pdf_path, dpi=dpi, first_page=1, last_page=1)
        return images[0] if images else None


def parse_quota(text):
    """Analizza una stringa di quota e estrae i componenti."""
    original = text
    text = text.strip()
    
    # Normalizza simboli diametro
    text = text.replace("Ø", "⌀").replace("ø", "⌀").replace("O/", "⌀").replace("0/", "⌀")
    
    result = {"simbolo": "", "nominale": "", "tol_plus": "", "tol_minus": "", "classe": ""}
    
    # Estrai simbolo iniziale (⌀, R, M, S, ecc.)
    m = re.match(r"^([⌀RMSsGg])\s*", text)
    if m:
        result["simbolo"] = m.group(1).upper()
        text = text[m.end():].strip()
    
    # Pattern 1: Numero + classe tolleranza ISO (es. "20H7", "20 H7", "10h6", "50g6")
    m = re.match(r"^([\d.,]+)\s*([A-Za-z]\d+)\s*$", text)
    if m:
        result["nominale"] = m.group(1).replace(",", ".")
        result["classe"] = m.group(2)
        return result
    
    # Pattern 2: ±tolleranza simmetrica (es. "20±0.1", "20 ± 0.1")
    m = re.match(r"^([\d.,]+)\s*[±]\s*([\d.,]+)\s*$", text)
    if m:
        result["nominale"] = m.group(1).replace(",", ".")
        tol = m.group(2).replace(",", ".")
        result["tol_plus"] = f"+{tol}"
        result["tol_minus"] = f"-{tol}"
        return result
    
    # Pattern 3: Tolleranze asimmetriche +X/-Y (es. "20+0.1-0.2", "20 +0.1 -0.2")
    m = re.match(r"^([\d.,]+)\s*\+\s*([\d.,]+)\s*[-–]\s*([\d.,]+)\s*$", text)
    if m:
        result["nominale"] = m.group(1).replace(",", ".")
        result["tol_plus"] = f"+{m.group(2).replace(',', '.')}"
        result["tol_minus"] = f"-{m.group(3).replace(',', '.')}"
        return result
    
    # Pattern 4: Tolleranze asimmetriche -Y/+X (ordine inverso)
    m = re.match(r"^([\d.,]+)\s*[-–]\s*([\d.,]+)\s*\+\s*([\d.,]+)\s*$", text)
    if m:
        result["nominale"] = m.group(1).replace(",", ".")
        result["tol_minus"] = f"-{m.group(2).replace(',', '.')}"
        result["tol_plus"] = f"+{m.group(3).replace(',', '.')}"
        return result
    
    # Pattern 5: Tolleranze su righe separate (es. "20\n+0.1\n-0.2" o "20 0.1 0.2")
    m = re.match(r"^([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*$", text)
    if m:
        result["nominale"] = m.group(1).replace(",", ".")
        # Assume prima positiva, seconda negativa
        result["tol_plus"] = f"+{m.group(2).replace(',', '.')}"
        result["tol_minus"] = f"-{m.group(3).replace(',', '.')}"
        return result
    
    # Pattern 6: Solo +tolleranza (es. "20+0.1")
    m = re.match(r"^([\d.,]+)\s*\+\s*([\d.,]+)\s*$", text)
    if m:
        result["nominale"] = m.group(1).replace(",", ".")
        result["tol_plus"] = f"+{m.group(2).replace(',', '.')}"
        return result
    
    # Pattern 7: Solo -tolleranza (es. "20-0.1")
    m = re.match(r"^([\d.,]+)\s*[-–]\s*([\d.,]+)\s*$", text)
    if m:
        result["nominale"] = m.group(1).replace(",", ".")
        result["tol_minus"] = f"-{m.group(2).replace(',', '.')}"
        return result
    
    # Pattern 8: Numero con spazi (es. "20 . 5" -> "20.5")
    cleaned = re.sub(r"(\d)\s*[.,]\s*(\d)", r"\1.\2", text)
    m = re.match(r"^([\d.]+)\s*$", cleaned)
    if m:
        result["nominale"] = m.group(1)
        return result
    
    # Pattern 9: Solo numero
    m = re.match(r"^([\d.,]+)\s*$", text)
    if m:
        result["nominale"] = m.group(1).replace(",", ".")
        return result
    
    # Fallback: tutto come nominale
    result["nominale"] = original.strip()
    return result


class ProgressDialog(tk.Toplevel):
    """Dialog con barra di progresso."""
    
    def __init__(self, parent, title="Elaborazione"):
        super().__init__(parent)
        self.title(title)
        self.transient(parent)
        self.resizable(False, False)
        
        # Centra sulla finestra padre
        self.geometry("350x100")
        x = parent.winfo_x() + (parent.winfo_width() - 350) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 100) // 2
        self.geometry(f"+{x}+{y}")
        
        self.label = tk.Label(self, text="Inizializzazione...")
        self.label.pack(pady=(15, 5))
        
        self.progress = ttk.Progressbar(self, length=300, mode='determinate')
        self.progress.pack(pady=10)
        
        self.protocol("WM_DELETE_WINDOW", lambda: None)  # Blocca chiusura
        self.grab_set()
    
    def update_progress(self, value, text=""):
        self.progress['value'] = value
        if text:
            self.label.config(text=text)
        self.update()


class PallinatoreApp(tk.Tk):
    """Applicazione principale."""
    
    # Dimensione massima per display (pixel sul lato lungo)
    DISPLAY_MAX_SIZE = 2000
    PALLINO_RADIUS = 12
    
    def __init__(self):
        super().__init__()
        
        self.title("Pallinatore Quote v6")
        self.geometry("1400x900")
        
        # Stato
        self.image_path = None
        self.original_size = (0, 0)  # Dimensioni REALI dell'immagine
        self.working_image = None    # Immagine ridimensionata per lavorare
        self.display_image = None    # Immagine per display (zoomata)
        self.tk_image = None
        self.zoom = 1.0
        self.image_scale = 1.0       # Fattore scala tra originale e working
        
        self.ocr_results = []
        self.pallini = []
        self.next_id = 1
        
        # Stato drag
        self.dragging = None
        self.drag_offset = (0, 0)
        
        self._build_ui()
    
    def _build_ui(self):
        # Toolbar
        toolbar = tk.Frame(self, bd=1, relief=tk.RAISED)
        toolbar.pack(side=tk.TOP, fill=tk.X)
        
        tk.Button(toolbar, text="📂 Apri", command=self.open_file).pack(side=tk.LEFT, padx=2, pady=2)
        tk.Button(toolbar, text="🔍 Scansiona OCR", command=self.scan_ocr).pack(side=tk.LEFT, padx=2, pady=2)
        tk.Button(toolbar, text="🎯 Auto Pallina", command=self.auto_pallina).pack(side=tk.LEFT, padx=2, pady=2)
        
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=5)
        
        tk.Button(toolbar, text="🔢 Rinumera", command=self.rinumera).pack(side=tk.LEFT, padx=2, pady=2)
        tk.Button(toolbar, text="🗑️ Pulisci tutto", command=self.clear_pallini).pack(side=tk.LEFT, padx=2, pady=2)
        
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=5)
        
        tk.Button(toolbar, text="📊 Esporta Excel", command=self.export_excel).pack(side=tk.LEFT, padx=2, pady=2)
        tk.Button(toolbar, text="🖼️ Esporta Immagine", command=self.export_image).pack(side=tk.LEFT, padx=2, pady=2)
        tk.Button(toolbar, text="📄 Esporta PDF", command=self.export_pdf).pack(side=tk.LEFT, padx=2, pady=2)
        
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=5)
        
        tk.Label(toolbar, text="Zoom:").pack(side=tk.LEFT, padx=2)
        tk.Button(toolbar, text="-", width=2, command=self.zoom_out).pack(side=tk.LEFT)
        self.zoom_label = tk.Label(toolbar, text="100%", width=5)
        self.zoom_label.pack(side=tk.LEFT)
        tk.Button(toolbar, text="+", width=2, command=self.zoom_in).pack(side=tk.LEFT)
        
        self.show_boxes_var = tk.BooleanVar(value=True)
        tk.Checkbutton(toolbar, text="Mostra box OCR", variable=self.show_boxes_var, 
                       command=self.redraw).pack(side=tk.LEFT, padx=10)
        
        # Istruzioni
        tk.Label(toolbar, text="│ Trascina=sposta │ DX=elimina │ Click=aggiungi", 
                 fg="gray").pack(side=tk.RIGHT, padx=10)
        
        # Area principale
        main = tk.PanedWindow(self, orient=tk.HORIZONTAL)
        main.pack(fill=tk.BOTH, expand=True)
        
        # Canvas con scrollbar
        canvas_frame = tk.Frame(main)
        main.add(canvas_frame, stretch="always")
        
        self.canvas = tk.Canvas(canvas_frame, bg="#404040", cursor="crosshair")
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        vscroll = tk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        vscroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        hscroll = tk.Scrollbar(self, orient=tk.HORIZONTAL, command=self.canvas.xview)
        hscroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.canvas.configure(xscrollcommand=hscroll.set, yscrollcommand=vscroll.set)
        
        # Eventi mouse
        self.canvas.bind("<Button-1>", self.on_mouse_down)
        self.canvas.bind("<B1-Motion>", self.on_mouse_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_mouse_up)
        self.canvas.bind("<Button-3>", self.on_right_click)
        
        # Tabella pallini
        table_frame = tk.Frame(main, width=350)
        main.add(table_frame)
        
        columns = ("id", "quota", "x", "y")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=30)
        self.tree.heading("id", text="ID")
        self.tree.heading("quota", text="Quota")
        self.tree.heading("x", text="X")
        self.tree.heading("y", text="Y")
        self.tree.column("id", width=40, anchor=tk.CENTER)
        self.tree.column("quota", width=180)
        self.tree.column("x", width=50, anchor=tk.CENTER)
        self.tree.column("y", width=50, anchor=tk.CENTER)
        
        tree_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Doppio click per eliminare dalla tabella
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        
        # Status bar
        self.status = tk.StringVar(value="Pronto. Apri un'immagine per iniziare.")
        tk.Label(self, textvariable=self.status, anchor=tk.W, relief=tk.SUNKEN).pack(side=tk.BOTTOM, fill=tk.X)
    
    # ============ FILE ============
    
    def open_file(self):
        path = filedialog.askopenfilename(
            title="Seleziona disegno",
            filetypes=[
                ("Immagini e PDF", "*.png *.jpg *.jpeg *.tif *.tiff *.bmp *.pdf"),
                ("Tutti i file", "*.*")
            ]
        )
        if not path:
            return
        
        try:
            import gc
            
            self.status.set("Caricamento...")
            self.update()
            
            # Libera memoria precedente
            if self.working_image:
                del self.working_image
            if self.display_image:
                del self.display_image
            gc.collect()
            
            # Carica immagine
            if path.lower().endswith(".pdf"):
                self.status.set("Conversione PDF...")
                self.update()
                img = pdf_to_image(path)
            else:
                img = Image.open(path)
            
            img = img.convert("RGB")
            
            # Salva dimensioni originali
            self.original_size = img.size
            orig_w, orig_h = img.size
            
            # Calcola se serve ridimensionamento per il display
            max_dim = max(orig_w, orig_h)
            if max_dim > self.DISPLAY_MAX_SIZE:
                self.image_scale = self.DISPLAY_MAX_SIZE / max_dim
                new_w = int(orig_w * self.image_scale)
                new_h = int(orig_h * self.image_scale)
                
                self.status.set(f"Ridimensionamento {orig_w}x{orig_h} → {new_w}x{new_h}...")
                self.update()
                
                self.working_image = img.resize((new_w, new_h), Image.LANCZOS)
                
                # Libera immagine originale
                del img
                gc.collect()
                
                print(f"[DEBUG] Immagine grande ridimensionata: {orig_w}x{orig_h} -> {new_w}x{new_h}")
                print(f"[DEBUG] Fattore scala display: {self.image_scale:.4f}")
            else:
                self.working_image = img
                self.image_scale = 1.0
            
            self.image_path = path
            self.zoom = 1.0
            self.ocr_results = []
            self.clear_pallini()
            
            self._update_display()
            
            size_str = f"{orig_w}x{orig_h}"
            if self.image_scale < 1.0:
                size_str += f" (display ridotto)"
            self.status.set(f"Caricato: {os.path.basename(path)} ({size_str})")
            
            gc.collect()
            
        except Exception as e:
            messagebox.showerror("Errore", f"Impossibile aprire il file:\n{e}")
    
    # ============ ZOOM ============
    
    def zoom_in(self):
        if self.working_image:
            self.zoom = min(3.0, self.zoom * 1.25)
            self._update_display()
    
    def zoom_out(self):
        if self.working_image:
            self.zoom = max(0.1, self.zoom / 1.25)
            self._update_display()
    
    def _update_display(self):
        if self.working_image is None:
            return
        
        import gc
        
        w = int(self.working_image.width * self.zoom)
        h = int(self.working_image.height * self.zoom)
        
        # Libera vecchia display_image
        if self.display_image is not None:
            del self.display_image
        
        self.display_image = self.working_image.resize((w, h), Image.LANCZOS)
        self.zoom_label.config(text=f"{int(self.zoom*100)}%")
        
        gc.collect()
        self.redraw()
    
    # ============ OCR ============
    
    # Dimensione massima per OCR (pixel sul lato lungo)
    OCR_MAX_SIZE = 2500
    
    def scan_ocr(self):
        if self.working_image is None:
            messagebox.showinfo("Info", "Carica prima un'immagine.")
            return
        
        # Mostra dialog progresso
        progress = ProgressDialog(self, "Scansione OCR")
        
        def update_progress(value, text):
            progress.update_progress(value, text)
        
        ocr_path = None
        converted_file = None
        
        try:
            import gc
            
            work_w, work_h = self.working_image.size
            orig_w, orig_h = self.original_size
            
            # Calcola se serve ulteriore ridimensionamento per OCR
            max_dim = max(work_w, work_h)
            if max_dim > self.OCR_MAX_SIZE:
                ocr_scale = self.OCR_MAX_SIZE / max_dim
                new_w = int(work_w * ocr_scale)
                new_h = int(work_h * ocr_scale)
                needs_resize = True
            else:
                ocr_scale = 1.0
                new_w, new_h = work_w, work_h
                needs_resize = False
            
            # Determina se serve conversione formato
            needs_conversion = (
                self.image_path is None or
                self.image_path.lower().endswith('.pdf') or
                self.image_path.lower().endswith('.tif') or
                self.image_path.lower().endswith('.tiff')
            )
            
            # Crea file per OCR
            if needs_resize or needs_conversion:
                update_progress(2, "Preparazione immagine per OCR...")
                
                if needs_resize:
                    ocr_image = self.working_image.resize((new_w, new_h), Image.LANCZOS)
                else:
                    ocr_image = self.working_image
                
                if self.image_path:
                    base = os.path.splitext(self.image_path)[0]
                    converted_file = f"{base}_ocr.png"
                else:
                    converted_file = os.path.join(os.path.expanduser("~"), "ocr_temp.png")
                
                ocr_image.save(converted_file, "PNG", optimize=True)
                ocr_path = converted_file
                
                if needs_resize:
                    del ocr_image
                    gc.collect()
                
                print(f"[DEBUG] File OCR: {new_w}x{new_h}")
            else:
                ocr_path = self.image_path
                ocr_scale = 1.0
            
            update_progress(5, "Avvio OCR...")
            
            # Esegui OCR
            self.ocr_results = run_ocr(ocr_path, update_progress)
            
            # IMPORTANTE: riscala le coordinate
            # Le coordinate OCR sono relative all'immagine OCR
            # Dobbiamo riportarle alle coordinate dell'immagine WORKING (non originale!)
            # perché il display usa working_image
            if ocr_scale != 1.0:
                inv_scale = 1.0 / ocr_scale
                for r in self.ocr_results:
                    r["box"] = r["box"] * inv_scale
                print(f"[DEBUG] Coordinate riscalate: {inv_scale:.4f}x")
            
            gc.collect()
            
            quote_count = sum(1 for r in self.ocr_results if re.search(r"\d", r["text"]))
            
            msg = f"OCR completato: {len(self.ocr_results)} testi, {quote_count} quote"
            self.status.set(msg)
            self.redraw()
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Errore OCR", str(e))
            self.status.set("Errore durante OCR")
        finally:
            progress.destroy()
            import gc
            gc.collect()
    
    # ============ PALLINI ============
    
    def auto_pallina(self):
        if not self.ocr_results:
            messagebox.showinfo("Info", "Esegui prima la scansione OCR.")
            return
        
        self.clear_pallini()
        
        for r in self.ocr_results:
            text = r["text"].strip()
            box = r["box"]
            
            if not re.search(r"\d", text):
                continue
            
            # Posizione: a sinistra del box, centrato verticalmente
            x = float(box[:, 0].min()) - 20
            y = float(box[:, 1].mean())
            x = max(15, x)
            
            self._add_pallino(x, y, text)
        
        self.redraw()
        self.status.set(f"Creati {len(self.pallini)} pallini (trascinabili)")
    
    def _add_pallino(self, x, y, text):
        pallino = {
            "id": self.next_id,
            "x": x,
            "y": y,
            "text": text
        }
        self.pallini.append(pallino)
        self.next_id += 1
        self._refresh_tree()
    
    def _remove_pallino(self, index):
        if 0 <= index < len(self.pallini):
            del self.pallini[index]
            self._refresh_tree()
            self.redraw()
    
    def _refresh_tree(self):
        """Aggiorna la tabella."""
        for item in self.tree.get_children():
            self.tree.delete(item)
        for p in self.pallini:
            self.tree.insert("", tk.END, values=(p["id"], p["text"], int(p["x"]), int(p["y"])))
    
    def clear_pallini(self):
        self.pallini = []
        self.next_id = 1
        self._refresh_tree()
        self.redraw()
    
    def rinumera(self):
        """Rinumera i pallini in ordine sequenziale."""
        if not self.pallini:
            return
        
        # Ordina per Y poi per X (dall'alto in basso, da sinistra a destra)
        self.pallini.sort(key=lambda p: (p["y"], p["x"]))
        
        # Rinumera
        for i, p in enumerate(self.pallini, 1):
            p["id"] = i
        
        self.next_id = len(self.pallini) + 1
        self._refresh_tree()
        self.redraw()
        self.status.set(f"Rinumerati {len(self.pallini)} pallini")
    
    # ============ MOUSE EVENTS ============
    
    def _find_pallino_at(self, canvas_x, canvas_y):
        """Trova pallino alle coordinate canvas. Ritorna indice o None."""
        img_x = canvas_x / self.zoom
        img_y = canvas_y / self.zoom
        
        for i, p in enumerate(self.pallini):
            dist = ((p["x"] - img_x)**2 + (p["y"] - img_y)**2)**0.5
            if dist < self.PALLINO_RADIUS / self.zoom + 5:
                return i
        return None
    
    def on_mouse_down(self, event):
        if self.working_image is None:
            return
        
        cx = self.canvas.canvasx(event.x)
        cy = self.canvas.canvasy(event.y)
        
        # Cerca pallino da trascinare
        idx = self._find_pallino_at(cx, cy)
        
        if idx is not None:
            # Inizia drag
            self.dragging = idx
            p = self.pallini[idx]
            self.drag_offset = (p["x"] - cx/self.zoom, p["y"] - cy/self.zoom)
            self.canvas.config(cursor="fleur")
        else:
            self.dragging = None
    
    def on_mouse_drag(self, event):
        if self.dragging is None:
            return
        
        cx = self.canvas.canvasx(event.x)
        cy = self.canvas.canvasy(event.y)
        
        # Aggiorna posizione pallino
        p = self.pallini[self.dragging]
        p["x"] = cx/self.zoom + self.drag_offset[0]
        p["y"] = cy/self.zoom + self.drag_offset[1]
        
        # Limita ai bordi (usa working_image)
        p["x"] = max(10, min(self.working_image.width - 10, p["x"]))
        p["y"] = max(10, min(self.working_image.height - 10, p["y"]))
        
        self.redraw()
    
    def on_mouse_up(self, event):
        if self.dragging is not None:
            self._refresh_tree()
            self.dragging = None
            self.canvas.config(cursor="crosshair")
        else:
            # Click semplice = aggiungi nuovo pallino
            cx = self.canvas.canvasx(event.x)
            cy = self.canvas.canvasy(event.y)
            
            img_x = cx / self.zoom
            img_y = cy / self.zoom
            
            # Cerca testo OCR vicino
            text = ""
            if self.ocr_results:
                best_dist = float('inf')
                for r in self.ocr_results:
                    box = r["box"]
                    bx = float(box[:, 0].mean())
                    by = float(box[:, 1].mean())
                    dist = ((bx - img_x)**2 + (by - img_y)**2)**0.5
                    if dist < best_dist and dist < 100:
                        best_dist = dist
                        text = r["text"]
            
            # Chiedi testo
            text = simpledialog.askstring(
                "Nuovo pallino",
                "Testo della quota:",
                initialvalue=text,
                parent=self
            )
            
            if text:
                self._add_pallino(img_x, img_y, text)
                self.redraw()
    
    def on_right_click(self, event):
        """Click destro = elimina pallino."""
        if not self.pallini:
            return
        
        cx = self.canvas.canvasx(event.x)
        cy = self.canvas.canvasy(event.y)
        
        idx = self._find_pallino_at(cx, cy)
        if idx is not None:
            self._remove_pallino(idx)
            self.status.set(f"Pallino eliminato. Rimasti: {len(self.pallini)}")
    
    def on_tree_double_click(self, event):
        """Doppio click su tabella = elimina."""
        sel = self.tree.selection()
        if not sel:
            return
        
        item = sel[0]
        values = self.tree.item(item, "values")
        pid = int(values[0])
        
        # Trova e rimuovi
        for i, p in enumerate(self.pallini):
            if p["id"] == pid:
                self._remove_pallino(i)
                break
    
    # ============ DISEGNO ============
    
    def redraw(self):
        self.canvas.delete("all")
        
        if self.display_image is None:
            return
        
        # Immagine
        self.tk_image = ImageTk.PhotoImage(self.display_image)
        self.canvas.create_image(0, 0, anchor=tk.NW, image=self.tk_image)
        self.canvas.configure(scrollregion=(0, 0, self.display_image.width, self.display_image.height))
        
        # Box OCR
        if self.show_boxes_var.get() and self.ocr_results:
            for r in self.ocr_results:
                box = r["box"]
                text = r["text"]
                
                if not re.search(r"\d", text):
                    continue
                
                x0 = float(box[:, 0].min()) * self.zoom
                y0 = float(box[:, 1].min()) * self.zoom
                x1 = float(box[:, 0].max()) * self.zoom
                y1 = float(box[:, 1].max()) * self.zoom
                
                self.canvas.create_rectangle(x0, y0, x1, y1, outline="yellow", width=1)
                self.canvas.create_text(x0, y0-2, text=text[:15], anchor=tk.SW, 
                                        fill="yellow", font=("Arial", 8))
        
        # Pallini
        for i, p in enumerate(self.pallini):
            x = p["x"] * self.zoom
            y = p["y"] * self.zoom
            r = self.PALLINO_RADIUS
            
            # Evidenzia pallino in drag
            if i == self.dragging:
                self.canvas.create_oval(x-r-2, y-r-2, x+r+2, y+r+2, outline="blue", width=2)
            
            # Cerchio
            self.canvas.create_oval(x-r, y-r, x+r, y+r, fill="white", outline="red", width=2)
            self.canvas.create_text(x, y, text=str(p["id"]), fill="red", font=("Arial", 9, "bold"))
    
    # ============ EXPORT ============
    
    def export_excel(self):
        if not self.pallini:
            messagebox.showinfo("Info", "Nessun pallino da esportare.")
            return
        
        path = filedialog.asksaveasfilename(
            title="Salva Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Quote"
            
            headers = ["ID", "Quota_raw", "Simbolo", "Nominale", "Tol+", "Tol-", "Classe"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            for row, p in enumerate(self.pallini, 2):
                parsed = parse_quota(p["text"])
                
                ws.cell(row=row, column=1, value=p["id"])
                ws.cell(row=row, column=2, value=p["text"])
                ws.cell(row=row, column=3, value=parsed["simbolo"])
                ws.cell(row=row, column=4, value=parsed["nominale"])
                ws.cell(row=row, column=5, value=parsed["tol_plus"])
                ws.cell(row=row, column=6, value=parsed["tol_minus"])
                ws.cell(row=row, column=7, value=parsed["classe"])
            
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 2
            
            wb.save(path)
            self.status.set(f"Esportato: {os.path.basename(path)}")
            messagebox.showinfo("Esportazione", f"File salvato:\n{path}")
            
        except ImportError:
            messagebox.showerror("Errore", "Installa openpyxl: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Errore", f"Errore esportazione:\n{e}")
    
    def _create_pallinated_image(self):
        """Crea un'immagine con i pallini disegnati sopra."""
        if self.working_image is None:
            return None
        
        from PIL import ImageDraw, ImageFont
        
        # Copia l'immagine di lavoro
        img = self.working_image.copy()
        draw = ImageDraw.Draw(img)
        
        # Cerca un font, fallback a default
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 16)
        except:
            try:
                font = ImageFont.truetype("arial.ttf", 16)
            except:
                font = ImageFont.load_default()
        
        # Disegna i pallini
        r = 15  # Raggio pallino
        for p in self.pallini:
            x, y = int(p["x"]), int(p["y"])
            
            # Cerchio bianco con bordo rosso
            draw.ellipse([x-r, y-r, x+r, y+r], fill="white", outline="red", width=3)
            
            # Numero centrato
            text = str(p["id"])
            bbox = draw.textbbox((0, 0), text, font=font)
            tw = bbox[2] - bbox[0]
            th = bbox[3] - bbox[1]
            draw.text((x - tw//2, y - th//2 - 2), text, fill="red", font=font)
        
        return img
    
    def export_image(self):
        """Esporta l'immagine con i pallini."""
        if self.working_image is None:
            messagebox.showinfo("Info", "Nessuna immagine caricata.")
            return
        
        path = filedialog.asksaveasfilename(
            title="Salva Immagine",
            defaultextension=".png",
            filetypes=[
                ("PNG", "*.png"),
                ("JPEG", "*.jpg *.jpeg"),
                ("TIFF", "*.tif *.tiff"),
                ("BMP", "*.bmp")
            ]
        )
        if not path:
            return
        
        try:
            img = self._create_pallinated_image()
            if img:
                # Per JPEG, converti in RGB se necessario
                if path.lower().endswith(('.jpg', '.jpeg')):
                    img = img.convert('RGB')
                img.save(path)
                self.status.set(f"Immagine salvata: {os.path.basename(path)}")
                messagebox.showinfo("Esportazione", f"Immagine salvata:\n{path}")
        except Exception as e:
            messagebox.showerror("Errore", f"Errore salvataggio immagine:\n{e}")
    
    def export_pdf(self):
        """Esporta l'immagine con i pallini in PDF."""
        if self.working_image is None:
            messagebox.showinfo("Info", "Nessuna immagine caricata.")
            return
        
        path = filedialog.asksaveasfilename(
            title="Salva PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")]
        )
        if not path:
            return
        
        try:
            img = self._create_pallinated_image()
            if img:
                # Converti in RGB per PDF
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                img.save(path, "PDF", resolution=100.0)
                self.status.set(f"PDF salvato: {os.path.basename(path)}")
                messagebox.showinfo("Esportazione", f"PDF salvato:\n{path}")
        except Exception as e:
            messagebox.showerror("Errore", f"Errore salvataggio PDF:\n{e}")


if __name__ == "__main__":
    app = PallinatoreApp()
    app.mainloop()
